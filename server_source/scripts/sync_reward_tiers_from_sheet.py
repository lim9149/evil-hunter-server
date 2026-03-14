# DEV-DIRECTION-LOCK: Portrait TownWorld UI / overlay panels / bottom fixed menu / visible hunt-return loop / original implementation only.
#!/usr/bin/env python3
"""Sync reward_tier(kind=worldboss|pvp) from the project Sheet into SQLite.

Supports 2 input formats:
1) Structured table in sheet '월드보스_PvP' with headers:
   - kind | rankMin | rankMax | multiplier
   (Korean aliases are also accepted: 종류/구분, 최소랭크, 최대랭크, 배율)

2) Fallback parse from the existing description cell that looks like:
   reward_tier(kind=worldboss): 1~1=1.0, 2~10=0.7, 11~50=0.4, 51~200=0.25, 201~∞=0.15

Usage:
  python scripts/sync_reward_tiers_from_sheet.py --xlsx EvilHunterTycoon_CompleteMasterTemplate_withPrompt.xlsx
  python scripts/sync_reward_tiers_from_sheet.py --xlsx ... --replace
  python scripts/sync_reward_tiers_from_sheet.py --xlsx ... --kinds worldboss pvp

Env:
  SQLITE_PATH : overrides DB path (default storage/evil_hunter.db)

Safety:
  - default mode: UPSERT only (does not delete existing tiers)
  - --replace : deletes existing tiers for kind first then inserts
"""

from __future__ import annotations

import sys
from pathlib import Path

# Ensure repo root is on PYTHONPATH when executed as a script
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import argparse
import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from storage.sqlite_db import upsert_reward_tier, bulk_replace_reward_tiers


_INFINITY_PATTERNS = ("∞", "inf", "infty", "infinite", "무한", "infinity")


def _norm(s: str) -> str:
    return str(s).strip()


def _is_infinity(v: object) -> bool:
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return False
    s = str(v).strip().lower()
    return any(p in s for p in _INFINITY_PATTERNS)


def _parse_tier_list(text: str) -> List[Tuple[int, Optional[int], float]]:
    # e.g. "1~1=1.0, 2~10=0.7, 201~∞=0.15"
    # allow "-" or "~" for range
    t = str(text)
    parts = re.split(r"[,\\n]", t)
    out: List[Tuple[int, Optional[int], float]] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.search(r"(\\d+)\\s*[~\\-]\\s*([^=]+?)\\s*=\\s*([0-9]*\\.?[0-9]+)", p)
        if not m:
            continue
        rmin = int(m.group(1))
        rmax_raw = m.group(2).strip()
        mul = float(m.group(3))
        rmax = None if _is_infinity(rmax_raw) else int(re.sub(r"[^0-9]", "", rmax_raw))
        out.append((rmin, rmax, mul))
    # sort & basic validate
    out.sort(key=lambda x: (x[0], 2147483647 if x[1] is None else x[1]))
    return out


def _try_parse_from_structured_table(wb) -> Dict[str, List[Tuple[int, Optional[int], float]]]:
    ws = wb["월드보스_PvP"]
    header_aliases = {
        "kind": {"kind", "종류", "구분", "타입"},
        "rankMin": {"rankmin", "rank_min", "최소랭크", "랭크min", "rankfrom"},
        "rankMax": {"rankmax", "rank_max", "최대랭크", "랭크max", "rankto"},
        "multiplier": {"multiplier", "배율", "보상배율", "랭크배율"},
    }

    def match_header(v: object) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip().lower()
        for key, aliases in header_aliases.items():
            if s in aliases:
                return key
        return None

    # find header row
    header_row = None
    header_map: Dict[str, int] = {}
    for r in range(1, min(ws.max_row, 200) + 1):
        tmp = {}
        for c in range(1, min(ws.max_column, 30) + 1):
            key = match_header(ws.cell(r, c).value)
            if key:
                tmp[key] = c
        if {"kind", "rankMin", "rankMax", "multiplier"}.issubset(set(tmp.keys())):
            header_row = r
            header_map = tmp
            break

    if header_row is None:
        return {}

    buckets: Dict[str, List[Tuple[int, Optional[int], float]]] = {"worldboss": [], "pvp": []}

    for r in range(header_row + 1, min(ws.max_row, header_row + 200) + 1):
        kind = ws.cell(r, header_map["kind"]).value
        if kind is None:
            continue
        k = str(kind).strip().lower()
        if k not in ("worldboss", "pvp"):
            continue
        rmin = ws.cell(r, header_map["rankMin"]).value
        rmax = ws.cell(r, header_map["rankMax"]).value
        mul = ws.cell(r, header_map["multiplier"]).value
        if rmin is None or mul is None:
            continue
        rrmin = int(rmin)
        rrmax = None if _is_infinity(rmax) else int(rmax)
        mmul = float(mul)
        buckets[k].append((rrmin, rrmax, mmul))

    for k in buckets:
        buckets[k].sort(key=lambda x: (x[0], 2147483647 if x[1] is None else x[1]))

    # If both empty, treat as not found
    if not buckets["worldboss"] and not buckets["pvp"]:
        return {}

    return buckets


def _parse_from_description_cells(wb) -> Dict[str, List[Tuple[int, Optional[int], float]]]:
    ws = wb["월드보스_PvP"]
    buckets: Dict[str, List[Tuple[int, Optional[int], float]]] = {}
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(r, c).value
            if not v or not isinstance(v, str):
                continue
            s = v.strip()
            m = re.search(r"reward_tier\\(kind\\s*=\\s*(worldboss|pvp)\\)\\s*:\\s*(.+)$", s, re.IGNORECASE)
            if m:
                kind = m.group(1).lower()
                tiers = _parse_tier_list(m.group(2))
                if tiers:
                    buckets[kind] = tiers
    return buckets


def load_tiers_from_xlsx(xlsx_path: str) -> Dict[str, List[Tuple[int, Optional[int], float]]]:
    wb = load_workbook(xlsx_path, data_only=True)

    buckets = _try_parse_from_structured_table(wb)
    if buckets:
        return buckets

    buckets = _parse_from_description_cells(wb)
    if buckets:
        return buckets

    raise RuntimeError("Could not find reward_tier data in sheet '월드보스_PvP'. Add a structured table or the description line.")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to EvilHunterTycoon_CompleteMasterTemplate_withPrompt.xlsx")
    ap.add_argument("--replace", action="store_true", help="Replace tiers per kind (delete then insert)")
    ap.add_argument("--kinds", nargs="*", default=["worldboss", "pvp"], help="Kinds to sync: worldboss pvp")
    args = ap.parse_args()

    xlsx = args.xlsx
    if not os.path.exists(xlsx):
        raise SystemExit(f"xlsx not found: {xlsx}")

    buckets = load_tiers_from_xlsx(xlsx)

    kinds = [k.lower() for k in args.kinds]
    for k in kinds:
        if k not in ("worldboss", "pvp"):
            raise SystemExit("--kinds must be worldboss and/or pvp")
        tiers = buckets.get(k, [])
        if not tiers:
            print(f"[skip] no tiers for kind={k}")
            continue

        if args.replace:
            res = bulk_replace_reward_tiers(k, tiers)
            print(f"[replace] kind={k} inserted={res['count']}")
        else:
            n = 0
            for rmin, rmax, mul in tiers:
                upsert_reward_tier(k, rmin, rmax, mul)
                n += 1
            print(f"[upsert] kind={k} upserted={n}")


if __name__ == "__main__":
    main()